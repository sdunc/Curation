# Stephen Duncanson
# a collection of functions to operate on spreadsheets

from openpyxl import load_workbook
import openpyxl
import waterp

def cols_to_tuples(name_of_file, col_tuple):
    """ 
    input:
    name_of_file - string. the name of the file to open
    cols - an array of which columns to grab data from
    if cols is None (default) all cols are grabbed.
    for each row y in x columns:
    return an array of length y where each element is
    (x0, x1) - The contents of each column.
    ex: 

    lat | lon
    ----------
     11 | -2
    .
    .
    .

    -> [(11,-2),...]
    """
    tuple_array = []

    wb = load_workbook(filename = name_of_file)
    ws = wb.active

    # row 1 is header: (latitude, longitude)
    for row_number in range(2, ws.max_row+1):
        tuple_array.append((ws.cell(row=row_number,column=5).value,
                            ws.cell(row=row_number,column=4).value))

    return tuple_array


if __name__ == "__main__":
    coord_array = cols_to_tuples("TGDR397_Plant_Accession.xlsx", [2,3])
    print(coord_array)

    # lets load the coastline polygon also
    # this should probably be split off into a nice function in waterp
    # anyway.
    
    for coord in coord_array:
        if waterp.point_in_water(coord):
            print(coord, " Is in water!!")
        else:
            #print(coord, "is not in water!")
            pass

